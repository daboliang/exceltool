# -*- coding: utf-8 -*-
import sys  # sys模块提供对Python解释器使用或维护的一些变量的访问，以及与解释器交互的函数
import os
from PyQt5 import QtWidgets  # QtWidgets指Qt控件模块，它提供了很多用于创建图形用户界面的工具，便于我们创建窗口
from PyQt5.QtGui import QIcon  # QIcon用于给界面设置logo
import xlrd
import datetime
import time
from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment


class MainGUI(QtWidgets.QMainWindow):
    '''
    文件目录框使用QLineEdit()、按钮使用QPushButton()、
    结果显示框使用QListWidget()，
    整体的布局使用QGridLayout()
    '''
    # 设置用户界面
    def __init__(self):
        super().__init__()
        self.setWindowTitle("领料明细汇总")  # 设置窗口标题
        self.resize(800, 400)  # 设置窗口大小，单位为像素
        self.main_widget = QtWidgets.QWidget()  # 创建界面
        self.main_widget_layout = QtWidgets.QGridLayout()  # 选取布局为网格布局(多行多列)
        self.main_widget.setLayout(self.main_widget_layout)  # 设置布局

        self.input = QtWidgets.QLineEdit()  # 文本编辑框1
        self.input_btn = QtWidgets.QPushButton("选择输入文件夹")  # 按钮1及名称
        self.output = QtWidgets.QLineEdit()  # 文本编辑框2
        self.output_btn = QtWidgets.QPushButton("选择输出文件夹")  # 按钮2及名称
        self.show_result = QtWidgets.QListWidget()  # 列表控件
        self.run = QtWidgets.QPushButton("执行汇总")  # 按钮3及名称

        self.main_widget_layout.addWidget(
            self.input, 0, 0, 1, 2)  # 文本编辑框1放在第1行1列，占用1行2列
        self.main_widget_layout.addWidget(
            self.input_btn, 0, 2, 1, 1)  # 按钮1放在第1行第3列，占用1行1列
        self.main_widget_layout.addWidget(self.output, 1, 0, 1, 2)
        self.main_widget_layout.addWidget(self.output_btn, 1, 2, 1, 1)
        self.main_widget_layout.addWidget(self.run, 2, 2, 1, 1)
        self.main_widget_layout.addWidget(self.show_result, 3, 0, 3, 3)

        self.setCentralWidget(self.main_widget)  # 设置QMainWindow的中心窗口

        # 将"选择输入文件夹"按钮绑定Choice_dir_input函数
        self.input_btn.clicked.connect(self.Choice_dir_input)
        # 将"选择输出文件夹"按钮绑定Choice_dir_output函数
        self.output_btn.clicked.connect(self.Choice_dir_output)
        self.run.clicked.connect(self.Summary_data)  # “执行汇总”按钮绑定Summary_data函数

    def Choice_dir_input(self):
        # 选择目录操作
        dir_path = QtWidgets.QFileDialog.getExistingDirectory(
            self, "请选择文件夹路径", "D:\\")
        # 将选择的目录显示在文本编辑框中
        self.input.setText(dir_path)

    def Choice_dir_output(self):
        dir_path = QtWidgets.QFileDialog.getExistingDirectory(
            self, "请选择文件夹路径", "D:\\")
        self.output.setText(dir_path)

    def Get_data(self, file):
        '''获取单个Excel文件中的资料'''
        wb = xlrd.open_workbook(file)
        ws = wb.sheets()[0]
        data = {}
        for row in range(7, ws.nrows - 2):
            card_id = ws.cell(2, 16).value
            car = ws.cell(3, 16).value
            dt = ws.cell(row, 0).value
            if isinstance(dt, float):
                date_time = xlrd.xldate.xldate_as_datetime(dt, 0)
            else:
                date_time = datetime.datetime.strptime(dt, '%Y-%m-%d %H:%M:%S')
            business = ws.cell(row, 2).value
            model = ws.cell(row, 3).value
            qty = ws.cell(row, 4).value
            unit_price = ws.cell(row, 6).value
            price = ws.cell(row, 8).value
            reward = ws.cell(row, 9).value
            discount = ws.cell(row, 11).value
            balance = ws.cell(row, 13).value
            location = str(ws.cell(row, 15).value).strip()
            operator = ws.cell(row, 17).value
            date = date_time.date()
            time = date_time.time()
            info_list = [
                card_id,
                car,
                date_time,
                business,
                model,
                qty,
                unit_price,
                price,
                reward,
                discount,
                balance,
                location,
                operator,
                date,
                time]
            data.setdefault(date, [])
            if info_list[3] != "备注":
                data[date].append(info_list)
        # 增加当日加油次数
        for key in data.keys():
            for i in data[key]:
                i.append(len(data[key]))
        return data

    def Get_file_path(self, path):
        files = []
        for file in os.listdir(path):
            if file.endswith(".xls"):  # 排除文件夹内的其它干扰文件
                files.append(path + "\\" + file)
        return files

    def Get_current_time(self):
        time_stamp = time.time()  # 当前时间的时间戳
        local_time = time.localtime(time_stamp)  #
        str_time = time.strftime('%Y-%m-%d %H.%M.%S', local_time)
        return str_time

    def Summary_data(self, files):
        thin = Side(border_style="thin", color="000000")  # 定义边框粗细及颜色
        title = ['部门', '部门编号', '时间', '业务类型', '品种', '数量', '单价', '金额', '额外值',
                 '调整', '剩余', '库位', '操作员', '领取日期', '领取时间', '领取次数']

        wb = Workbook()
        ws = wb.active
        ws.merge_cells("A1:P1")
        ws.cell(1, 1).value = "领料明细汇总表"
        ws.cell(1, 1).font = Font(name=u'黑体', bold=True, size=18)
        ws.row_dimensions[1].height = 22.2
        ws.cell(
            1,
            1).alignment = Alignment(
            horizontal="center",
            vertical="center")
        ws.append(title)

        # 插入数据
        # 获取文本编辑框中的输入文件目录，并获取目录下的xls文件
        files = self.Get_file_path(self.input.text())
        for file in files:
            data = self.Get_data(file)
            for key in data.keys():
                for i in data[key]:
                    ws.append(i)
            f = QtWidgets.QListWidgetItem(f"{file} 的内容已加入总表.")  # 创建一个显示项
            self.show_result.addItem(f)  # 将结果添加到部件中

        # 设置字号，对齐，缩小字体填充，加边框
        # Font(bold=True)可加粗字体
        for row_number in range(2, ws.max_row + 1):
            for col_number in range(1, 17):
                c = ws.cell(row=row_number, column=col_number)
                c.font = Font(size=9)
                c.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                c.alignment = Alignment(horizontal="left", vertical="center")

        col_name = list("ABCDEFGHIJKLMNOP")
        col_width = [8, 8, 16, 8, 16, 8, 8, 9.8, 8, 8, 8, 11, 8.3, 9, 8, 8]
        for i in range(len(col_name)):
            ws.column_dimensions[col_name[i]].width = col_width[i]

        ws.column_dimensions.group('I', 'K', hidden=True)
        ws.column_dimensions.group('N', 'O', hidden=True)

        wb.save(f"{self.output.text()}\\领料明细汇总表{self.Get_current_time()}.xlsx")
        f = QtWidgets.QListWidgetItem(
            f"\n领料明细汇总表{self.Get_current_time()}.xlsx 已生成，请去输出文件夹查看.")  # 创建一个显示项
        self.show_result.addItem(f)  # 将结果添加到部件中


def main():
    # [固定写法]实例化应用，sys.argv是一个从程序外部获取参数的桥梁
    app = QtWidgets.QApplication(sys.argv)
    app.setWindowIcon(QIcon("PO.ico"))  # 设置界面左上角图标
    gui = MainGUI()  # [固定写法]调用MainGUI类
    gui.show()  # [固定写法]显示窗口
    sys.exit(app.exec_())  # [固定写法]调用exec_()方法运行程序。sys.exit()用于程序的正常退出


if __name__ == '__main__':
    main()
"""
@Time        : 2021/10/5
@Author      : 89464
@File        : 文件
@Description :
"""
